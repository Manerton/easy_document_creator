from openpyxl import workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet import worksheet

from token_xlsx import TokenXlsx, TokenTypeXlsx, TokenTypeCollection


class XlsxAnalyzer:
    __document: workbook
    active_worksheet: worksheet
    __tokens_xlsx: TokenXlsx
    __data: any
    __used_cell: list

    def __init__(self):
        self.__document = None
        self.__tokens_xlsx = TokenXlsx()
        self.__used_cell = []

    def open_document(self, name_doc: str):
        self.__document = load_workbook(name_doc)

    def start_search(self):
        worksheets = self.__document.worksheets
        for worksheet in worksheets:
            self.active_worksheet = worksheet
            for row in worksheet.iter_rows():
                for cell in row:
                    self.__search_token(cell)

    # Поиск токена в ячейке
    def __search_token(self, cell: Cell):
        cell_text = cell.value
        if cell_text is None:
            return
        if "{{" and "}}" in cell_text:
            self.__analyze_type_token(cell)

    # Анализ типа токена
    def __analyze_type_token(self, cell: Cell):
        for word in cell.value.split():
            if "{{" and "}}" in word:
                if "." in word:
                    self.__tokens_xlsx.add_token_type_collection(word, word, cell)
                else:
                    self.__tokens_xlsx.add_token_str(cell)

    def init_data(self, data):
        self.__data = data

    # Замена токена в тексте ячейки на значение
    def __replace_str_in_cell(self, cell: Cell, full_token, new_name):
        cell.value = cell.value.replace(full_token, new_name)

    def __replace_str(self, key, value):
        temp_token: TokenTypeXlsx = self.__tokens_xlsx.tokens_str.get(key)
        if temp_token is None:
            return
        token_name = temp_token.token_name
        temp_token.cell.value = temp_token.cell.value.replace(token_name, value)

    def __isCellNotUsed(self, cell_name):
        for name in self.__used_cell:
            if cell_name == name:
                return False
        return True


    # !!!!!!!!Если ячейка merge что будет?
    def __add_line(self, cell: Cell, text):
        # Если строка ранее не добавлялась, то добавить
        if self.__isCellNotUsed(cell.row):
            self.__used_cell.append(cell.row)
            self.active_worksheet.insert_rows(cell.row+1)
        new_cell = self.active_worksheet.cell(cell.row+1, cell.column)
        new_cell.value = text
        return new_cell

    def __add_line_after_row(self, num_row):
        if self.__isCellNotUsed(num_row):
            self.__used_cell.append(num_row)
            self.active_worksheet.insert_rows(num_row+1)
        return num_row + 1

    # def append_new_line_for_table(self, collection: TokenTypeCollection):
    #     for sub_token in collection:
    #         value: TokenTypeXlsx or TokenTypeCollection = collection.get(sub_token)
    #         type_value = type(value)
    #         if type_value == TokenTypeXlsx:
    #             self.add_line(value.cell)

    def __create_full_name(self, first_name, second_name):
        return first_name + '.' + second_name

    def __replace_dict(self, main_token: str, data, collection=None, be_more=False):
        last_token = any
        temp_token_collection = collection
        if collection is None:
            temp_token_collection: TokenTypeCollection = self.__tokens_xlsx.tokens_collection.get(main_token)

        if temp_token_collection is None:
            return

        for key in data:
            i = 1
            value = data.get(key)
            type_value = type(value)
            if type_value == str or type_value == int:
                sub_token: TokenTypeXlsx = temp_token_collection.sub_tokens.get(key)
                if sub_token is None:
                    continue
                new_cell: Cell = None
                if be_more:
                    new_cell = self.__add_line(sub_token.cell, sub_token.text_cell)
                self.__replace_str_in_cell(sub_token.cell, sub_token.token_name, str(value))
                if new_cell is not None:
                    sub_token.cell = new_cell
                last_token = sub_token
            elif type_value == dict:
                sub_token: TokenTypeCollection = temp_token_collection.sub_tokens.get(key)
                last_token = self.__replace_dict(key, value, sub_token)
            elif type_value == list:
                temp_type = type(value[0])
                if temp_type == dict:
                    sub_token: TokenTypeCollection = temp_token_collection.sub_tokens.get(key)
                    if sub_token is None:
                        continue
                    last_token = self.__replace_list_dict(key, value, sub_token)
                elif temp_type == str:
                    sub_token: TokenTypeXlsx = temp_token_collection.sub_tokens.get(key)
                    if sub_token is None:
                        continue
                    last_token = self.__replace_list(key, value, sub_token)
        return last_token

    def __min_row(self, collection, min=None):
        if min:
            min_glob = min
        else:
            min_glob = None
        for key in collection.sub_tokens:
            value = collection.sub_tokens.get(key)

            if value is None:
                return

            type_value = type(value)
            if type_value == TokenTypeXlsx:
                if min_glob is None or min_glob < value.num_row:
                    min_glob = value.num_row
            else:
                self.__min_row(value, min_glob)
        return min_glob

    def __restoring_collection(self, collection, start_row, min_row=None):
        if min_row is None:
            min_row = self.__min_row(collection)
        for key in collection.sub_tokens:
            value = collection.sub_tokens.get(key)

            if value is None:
                return

            type_value = type(value)
            if type_value == TokenTypeXlsx:
                last_row = start_row
                count = value.num_row - min_row
                for i in range(count):
                    last_row = self.__add_line_after_row(last_row)
                if value.merge:
                    new_min = last_row
                    for i in range(value.merge.end_row - value.merge.start_row):
                        last_row = self.__add_line_after_row(last_row)
                    new_max = last_row
                    value.merge.start_row = new_min
                    value.merge.end_row = new_max
                    value.cell = self.active_worksheet.cell(value.merge.start_row, value.merge.start_column)
                    value.cell.value = value.text_cell
                    value.cell.font = value.my_style.font.copy()
                    value.cell.alignment = value.my_style.alignment.copy()
                    self.active_worksheet.merge_cells(start_row=value.merge.start_row, start_column=value.merge.start_column,
                                                      end_row=value.merge.end_row, end_column=value.merge.end_column)
                else:
                    value.cell = self.active_worksheet.cell(last_row, value.cell.column)
                    value.cell.value = value.text_cell
            else:
                self.__restoring_collection(value, start_row, min_row)

    def __all_same(self, data):
        for key in data:
            value = data.get(key)
            now_type = type(value)
            if str != now_type:
                return False
        return True

    def __replace_list_dict(self, main_key, list_data, collection=None):
        i = 1
        temp_token_collection = collection
        if collection is None:
            temp_token_collection: TokenTypeCollection = self.__tokens_xlsx.tokens_collection.get(main_key)
        # Если коллекция не найдена, то пропустить её
        if temp_token_collection is None:
            return

        last_token = None
        simple_repeat = self.__all_same(list_data[0])
        for data in list_data:
            if simple_repeat and i < len(list_data):
                be_more = True
            else:
                be_more = False
            last_token = self.__replace_dict(main_key, data, temp_token_collection, be_more)
            if not simple_repeat and i < len(list_data):
                last_row = self.__add_line_after_row(last_token.cell.row)
                self.__restoring_collection(temp_token_collection, last_row)
            i = i + 1
        return last_token

    def __replace_list(self, key, data, collection=None):
        i = 1
        temp_token = collection
        if collection is None:
            temp_token: TokenTypeXlsx = self.__tokens_xlsx.tokens_str.get(key)

        if temp_token is None:
            return

        for value in data:
            new_cell = None
            if i < len(data):
                i = i+1
                new_cell = self.__add_line(temp_token.cell, temp_token.text_cell)
            temp_token.cell.value = temp_token.cell.value.replace(temp_token.token_name, value)
            if new_cell:
                temp_token.cell = new_cell
        return temp_token

    def __analyze_type_in_list(self, main_key, list_data):
        temp_type = type(list_data[0])
        if temp_type == dict:
            self.__replace_list_dict(main_key, list_data)
        elif temp_type == str:
            self.__replace_list(main_key, list_data)

    def start_replace(self):
        for key in self.__data:
            value = self.__data.get(key)
            type_value = type(value)
            if type_value == str:
                self.__replace_str(key, value)
            if type_value == list:
                self.__analyze_type_in_list(key, value)
            if type_value == dict:
                self.__replace_dict(key, value)

    def get_token(self):
        return self.__tokens_xlsx.tokens_str

    def save_document(self, new_name: str):
        self.__document.save(new_name)
